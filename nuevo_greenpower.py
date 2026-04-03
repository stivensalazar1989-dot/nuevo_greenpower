"""
GreenPower - Gestión Energía Renovable
Script unificado con integración MySQL + Pillow (imágenes)
Requiere: pip install mysql-connector-python Pillow openpyxl reportlab
"""
import os

print("RUTA ACTUAL:", os.getcwd())
print("ARCHIVOS:", os.listdir())
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import List, Dict, Any, Optional
import json
import os
import re
from datetime import datetime
from typing import Optional, Tuple
from PIL import Image, ImageTk
import shutil
import tkinter as tk
from tkinter import ttk, messagebox
import ttkbootstrap as ttk


# ══════════════════════════════════════════════════════════════════
# MÓDULO DE IMÁGENES (Pillow)
# ══════════════════════════════════════════════════════════════════
from utils_imagen import (
    PILLOW_DISPONIBLE,
    validar_imagen,
    procesar_y_guardar,
    generar_miniatura,
    info_imagen,
    inicializar_carpetas,
)

inicializar_carpetas()

# ══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN BASE DE DATOS
# ══════════════════════════════════════════════════════════════════
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "database": "greenpower",
    "port": 3306,
}

try:
    import mysql.connector

    MYSQL_DISPONIBLE = True
except ImportError:
    MYSQL_DISPONIBLE = False


# ══════════════════════════════════════════════════════════════════
# CAPA DE BASE DE DATOS
# ══════════════════════════════════════════════════════════════════

class Database:
    def __init__(self):
        self.conn = None
        self.cursor = None
        self.conectado = False

    def conectar(self) -> bool:
        if not MYSQL_DISPONIBLE:
            return False
        try:
            self.conn = mysql.connector.connect(**DB_CONFIG)
            self.cursor = self.conn.cursor(dictionary=True)
            self.conectado = True
            return True
        except Exception:
            self.conectado = False
            return False

    def desconectar(self):
        try:
            if self.cursor: self.cursor.close()
            if self.conn:   self.conn.close()
        except Exception:
            pass
        self.conectado = False

    def ejecutar(self, query: str, params=None) -> bool:
        try:
            self.cursor.execute(query, params or ())
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            raise e

    def consultar(self, query: str, params=None) -> List[Dict]:
        try:
            self.cursor.execute(query, params or ())
            return self.cursor.fetchall()
        except Exception as e:
            raise e

    def consultar_uno(self, query: str, params=None) -> Optional[Dict]:
        try:
            self.cursor.execute(query, params or ())
            return self.cursor.fetchone()
        except Exception as e:
            raise e


db = Database()

plantas_mem: List[Dict] = []
equipos_mem: List[Dict] = []
estaciones_mem: List[Dict] = []


# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════

def _float(v) -> float:
    try:
        return float(v or 0)
    except:
        return 0.0


def _int(v) -> int:
    try:
        return int(float(v or 0))
    except:
        return 0


def _lat(coords: str) -> Optional[float]:
    try:
        return float(coords.split(',')[0].strip())
    except:
        return None


def _lon(coords: str) -> Optional[float]:
    try:
        return float(coords.split(',')[1].strip())
    except:
        return None


# ══════════════════════════════════════════════════════════════════
# VALIDADOR
# ══════════════════════════════════════════════════════════════════

class Validador:
    @staticmethod
    def es_numero(valor: str) -> bool:
        if not valor.strip(): return False
        return bool(re.match(r'^\d*\.?\d+$', valor.strip()))

    @staticmethod
    def codigo_valido(codigo: str) -> bool:
        codigo = codigo.strip().upper()
        return 3 <= len(codigo) <= 20 and bool(re.match(r'^[A-Z0-9]+$', codigo))

    @staticmethod
    def texto_valido(texto: str, min_len=1, max_len=100) -> bool:
        texto = texto.strip()
        return min_len <= len(texto) <= max_len and bool(
            re.match(r'^[a-zA-Z0-9\s\.,\-_áéíóúñÁÉÍÓÚÑ]+$', texto))

    @staticmethod
    def es_fecha_manual(fecha: str) -> bool:
        if not fecha or fecha == "AAAA-MM-DD": return False
        try:
            datetime.strptime(fecha, '%Y-%m-%d')
            return True
        except:
            return False


# ══════════════════════════════════════════════════════════════════
# EXPORTADOR
# ══════════════════════════════════════════════════════════════════

class Exportador:
    @staticmethod
    def exportar_excel(datos: List[Dict], titulo: str, nombre_archivo: str):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = titulo

            if not datos:
                ws['A1'] = "Sin datos para mostrar"
                wb.save(nombre_archivo)
                return True

            headers = list(datos[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for row_idx, fila in enumerate(datos, 3):
                for col_idx, valor in enumerate(fila.values(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(valor))

            ws.merge_cells(f'A1:{chr(64 + len(headers))}1')
            titulo_cell = ws['A1']
            titulo_cell.value = f"📊 {titulo} - GreenPower ({len(datos)} registros)"
            titulo_cell.font = Font(bold=True, size=14, color="2C3E50")
            titulo_cell.alignment = Alignment(horizontal="center")

            for col_num in range(1, len(headers) + 1):
                col_letter = chr(64 + col_num)
                max_length = 0
                for row in range(1, len(datos) + 3):
                    cell_value = str(ws.cell(row=row, column=col_num).value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                ws.column_dimensions[col_letter].width = min(max_length + 2, 25)

            wb.save(nombre_archivo)
            return True
        except Exception as e:
            print(f"❌ Error Excel: {e}")
            return False

    @staticmethod
    def exportar_pdf(datos: List[Dict], titulo: str, nombre_archivo: str):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.units import inch

            doc = SimpleDocTemplate(nombre_archivo, pagesize=A4,
                                    rightMargin=30, leftMargin=30,
                                    topMargin=60, bottomMargin=30)
            styles = getSampleStyleSheet()
            story = []

            story.append(Paragraph(f"<b>🌿 GreenPower - {titulo}</b>", styles['Title']))
            story.append(Spacer(1, 20))

            if datos:
                headers = list(datos[0].keys())
                data_table = [headers] + [[str(v) for v in row.values()] for row in datos[:20]]
                tabla = Table(data_table, colWidths=[0.8 * inch] * len(headers))
                tabla.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(tabla)
            else:
                story.append(Paragraph("<b>Sin datos</b>", styles['Normal']))

            doc.build(story)
            return True
        except Exception as e:
            print(f"❌ Error PDF: {e}")
            return False


# ══════════════════════════════════════════════════════════════════
# DIÁLOGO FILTROS EXPORTACIÓN
# ══════════════════════════════════════════════════════════════════

class DialogoFiltros:
    def __init__(self, parent, tipo: str):
        self.resultado = None
        self.tipo = tipo

        self.ventana = tk.Toplevel(parent)
        self.ventana.title(f"🔍 Filtros {tipo.title()}")
        self.ventana.geometry("400x300")
        self.ventana.resizable(False, False)
        self.ventana.transient(parent)
        self.ventana.grab_set()

        self._crear()

    def _crear(self):
        frame = ttk.Frame(self.ventana, padding=20)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="Filtrar por:", font=('Arial', 10, 'bold')).pack(pady=10)
        self.filtro_tipo = ttk.Combobox(frame,
                                        values=['TODOS', 'Operativos', 'No operativos', 'Solar', 'Eólica'],
                                        state='readonly', width=30)
        self.filtro_tipo.set('TODOS')
        self.filtro_tipo.pack(pady=5)

        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=15)
        ttk.Label(frame, text="Rango fechas (opcional):").pack()

        fecha_frame = ttk.Frame(frame)
        fecha_frame.pack(pady=5)
        ttk.Label(fecha_frame, text="Desde:").pack(side='left', padx=5)
        self.desde_fecha = ttk.Entry(fecha_frame, width=12)
        self.desde_fecha.insert(0, "AAAA-MM-DD")
        self.desde_fecha.pack(side='left', padx=5)
        ttk.Label(fecha_frame, text="Hasta:").pack(side='left', padx=5)
        self.hasta_fecha = ttk.Entry(fecha_frame, width=12)
        self.hasta_fecha.insert(0, "AAAA-MM-DD")
        self.hasta_fecha.pack(side='left', padx=5)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="✅ Exportar", command=self.aplicar).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="❌ Cancelar", command=self.ventana.destroy).pack(side='left', padx=10)

    def aplicar(self):
        self.resultado = {
            'tipo_filtro': self.filtro_tipo.get(),
            'fecha_desde': self.desde_fecha.get() if self.desde_fecha.get() != "AAAA-MM-DD" else None,
            'fecha_hasta': self.hasta_fecha.get() if self.hasta_fecha.get() != "AAAA-MM-DD" else None,
        }
        self.ventana.destroy()


# ══════════════════════════════════════════════════════════════════
# WIDGET REUTILIZABLE — SELECTOR DE IMAGEN CON PILLOW
# ══════════════════════════════════════════════════════════════════

class SelectorImagen(ttk.LabelFrame):
    def __init__(self, parent, titulo="🖼️ Imagen", categoria="planta", **kwargs):
        # quita `padding=10` (no lo soporta bien en tu contexto)
        super().__init__(parent, text=titulo, **kwargs)
        self.categoria = categoria
        self._ruta_origen = ""
        self._ruta_guardada = ""
        self._photo_ref = None
        self._construir()

    # --- Métodos de la clase SelectorImagen (no dentro de __init__) ---
    def _seleccionar(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar imagen",
            filetypes=[
                ("Imágenes permitidas", "*.jpg *.jpeg *.png *.gif"),
                ("JPEG", "*.jpg *.jpeg"),
                ("PNG", "*.png"),
                ("GIF", "*.gif"),
                ("Todos los archivos", "*.*"),
            ]
        )
        if not ruta:
            return
        resultado = validar_imagen(ruta)
        if not resultado.ok:
            self.lbl_val.config(text=f"❌ {resultado.mensaje}", foreground='red')
            return
        self._ruta_origen = ruta
        nombre = os.path.basename(ruta)
        self.lbl_archivo.config(text=nombre[:45] + ("…" if len(nombre) > 45 else ""),
                                foreground='black')
        self.lbl_val.config(text="✅ Imagen válida", foreground='green')
        self.lbl_info.config(text=info_imagen(ruta))
        thumb = generar_miniatura(ruta)
        if thumb:
            self._photo_ref = thumb
            self.canvas_thumb.config(image=thumb, text="", bg='white')

    def _construir(self):
        fila1 = ttk.Frame(self)
        fila1.pack(fill='x')
        ttk.Button(fila1, text="📂 Seleccionar imagen", command=self._seleccionar).pack(side='left', padx=(0, 10))
        self.lbl_archivo = ttk.Label(fila1, text="Ningún archivo seleccionado",
                                     foreground='gray', font=('Arial', 8))
        self.lbl_archivo.pack(side='left', fill='x', expand=True)
        fila2 = ttk.Frame(self)
        fila2.pack(fill='x', pady=8)
        self.canvas_thumb = tk.Label(fila2, text="Sin imagen",
                                     bg='#ecf0f1', width=14, height=7,
                                     relief='sunken', font=('Arial', 8),
                                     foreground='#95a5a6')
        self.canvas_thumb.pack(side='left', padx=(0, 12))
        opciones = ttk.Frame(fila2)
        opciones.pack(side='left', fill='y')
        ttk.Label(opciones, text="Convertir a:", font=('Arial', 8)).grid(row=0, column=0, sticky='w')
        self.cb_formato = ttk.Combobox(opciones, width=8, state='readonly',
                                       values=['Original', 'JPG', 'PNG'])
        self.cb_formato.set('Original')
        self.cb_formato.grid(row=0, column=1, padx=6, pady=2)
        ttk.Label(opciones, text="Filtro:", font=('Arial', 8)).grid(row=1, column=0, sticky='w')
        self.cb_filtro = ttk.Combobox(opciones, width=8, state='readonly',
                                      values=['Ninguno', 'Escala grises', 'Nitidez'])
        self.cb_filtro.set('Ninguno')
        self.cb_filtro.grid(row=1, column=1, padx=6, pady=2)
        self.lbl_info = ttk.Label(opciones, text="", foreground='#2980b9',
                                  font=('Arial', 8))
        self.lbl_info.grid(row=2, column=0, columnspan=2, sticky='w', pady=(4, 0))
        self.lbl_val = ttk.Label(fila2, text="", foreground='green',
                                 font=('Arial', 8), wraplength=160)
        self.lbl_val.pack(side='left', padx=8, anchor='n')

    def get_fname(self):
        """Devuelve la ruta de la imagen seleccionada"""
        return self._ruta_origen

    def ruta_actual(self):
        """Devuelve la ruta que usaría guardar/exportar"""
        return self._ruta_guardada or self._ruta_origen

    def limpiar(self):
        self._ruta_origen = ""
        self._ruta_guardada = ""
        self._photo_ref = None
        self.lbl_archivo.config(text="Ningún archivo seleccionado", foreground='gray')
        self.lbl_val.config(text="")
        self.lbl_info.config(text="")
        self.canvas_thumb.config(image='', text="Sin imagen",
                                 bg='#ecf0f1', foreground='#95a5a6')
        self.cb_formato.set('Original')
        self.cb_filtro.set('Ninguno')

        def ruta_actual(self):
            """Devuelve la ruta que usaría guardar/exportar"""
            return self._ruta_guardada or self._ruta_origen

    def _construir(self):
        # ── Fila superior: botón + info ─────────────────────────
        fila1 = ttk.Frame(self)
        fila1.pack(fill='x')

        ttk.Button(fila1, text="📂 Seleccionar imagen",
                   command=self._seleccionar).pack(side='left', padx=(0, 10))

        self.lbl_archivo = ttk.Label(fila1, text="Ningún archivo seleccionado",
                                     foreground='gray', font=('Arial', 8))
        self.lbl_archivo.pack(side='left', fill='x', expand=True)

        # ── Fila central: miniatura + opciones ──────────────────
        fila2 = ttk.Frame(self)
        fila2.pack(fill='x', pady=8)

        # Miniatura
        self.canvas_thumb = tk.Label(fila2, text="Sin imagen",
                                     bg='#ecf0f1', width=14, height=7,
                                     relief='sunken', font=('Arial', 8),
                                     foreground='#95a5a6')
        self.canvas_thumb.pack(side='left', padx=(0, 12))

        # Opciones Pillow
        opciones = ttk.Frame(fila2)
        opciones.pack(side='left', fill='y')

        ttk.Label(opciones, text="Convertir a:", font=('Arial', 8)).grid(row=0, column=0, sticky='w')
        self.cb_formato = ttk.Combobox(opciones, width=8, state='readonly',
                                       values=['Original', 'JPG', 'PNG'])
        self.cb_formato.set('Original')
        self.cb_formato.grid(row=0, column=1, padx=6, pady=2)

        ttk.Label(opciones, text="Filtro:", font=('Arial', 8)).grid(row=1, column=0, sticky='w')
        self.cb_filtro = ttk.Combobox(opciones, width=8, state='readonly',
                                      values=['Ninguno', 'Escala grises', 'Nitidez'])
        self.cb_filtro.set('Ninguno')
        self.cb_filtro.grid(row=1, column=1, padx=6, pady=2)

        # Info Pillow
        self.lbl_info = ttk.Label(opciones, text="", foreground='#2980b9',
                                  font=('Arial', 8))
        self.lbl_info.grid(row=2, column=0, columnspan=2, sticky='w', pady=(4, 0))

        # Validación
        self.lbl_val = ttk.Label(fila2, text="", foreground='green',
                                 font=('Arial', 8), wraplength=160)
        self.lbl_val.pack(side='left', padx=8, anchor='n')

    # ── Acciones ─────────────────────────────────────────────────

    def _seleccionar(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar imagen",
            filetypes=[
                ("Imágenes permitidas", "*.jpg *.jpeg *.png *.gif"),
                ("JPEG", "*.jpg *.jpeg"),
                ("PNG", "*.png"),
                ("GIF", "*.gif"),
                ("Todos los archivos", "*.*"),
            ]
        )
        if not ruta:
            return

        # Validar con Pillow
        resultado = validar_imagen(ruta)
        if not resultado.ok:
            self.lbl_val.config(text=f"❌ {resultado.mensaje}", foreground='red')
            return

        self._ruta_origen = ruta
        nombre = os.path.basename(ruta)
        self.lbl_archivo.config(text=nombre[:45] + ("…" if len(nombre) > 45 else ""),
                                foreground='black')
        self.lbl_val.config(text="✅ Imagen válida", foreground='green')
        self.lbl_info.config(text=info_imagen(ruta))

        # Miniatura
        thumb = generar_miniatura(ruta)
        if thumb:
            self._photo_ref = thumb
            self.canvas_thumb.config(image=thumb, text="", bg='white')

    def procesar(self, codigo: str) -> Optional[Tuple[bool, str]]:
        """
        Llama a procesar_y_guardar con las opciones elegidas.
        Retorna (True, ruta) o (False, mensaje).
        """
        if not self._ruta_origen:
            return True, ""  # Sin imagen seleccionada → no es error

        fmt_map = {'Original': None, 'JPG': 'JPEG', 'PNG': 'PNG'}
        filtro_map = {'Ninguno': None, 'Escala grises': 'gris', 'Nitidez': 'nitidez'}

        fmt = fmt_map.get(self.cb_formato.get())
        filtro = filtro_map.get(self.cb_filtro.get())

        ok, ruta = procesar_y_guardar(
            self._ruta_origen, codigo, self.categoria,
            convertir_a=fmt, filtro=filtro
        )
        if ok:
            self._ruta_guardada = ruta
        return ok, ruta

    def limpiar(self):
        self._ruta_origen = ""
        self._ruta_guardada = ""
        self._photo_ref = None
        self.lbl_archivo.config(text="Ningún archivo seleccionado", foreground='gray')
        self.lbl_val.config(text="")
        self.lbl_info.config(text="")
        self.canvas_thumb.config(image='', text="Sin imagen",
                                 bg='#ecf0f1', foreground='#95a5a6')
        self.cb_formato.set('Original')
        self.cb_filtro.set('Ninguno')

    @property
    def ruta_guardada(self):
        return self._ruta_guardada


# Alias para el type hint informal
def Tuple_o_None(*_): pass


# ══════════════════════════════════════════════════════════════════
# CRUD — PLANTAS
# ══════════════════════════════════════════════════════════════════

def agregar_planta(data: Dict[str, Any]) -> bool:
    codigo = data.get('codigo_planta', '').strip()
    if not codigo:
        return False
    if db.conectado:
        if db.consultar_uno("SELECT 1 FROM plantas WHERE codigo_planta=%s", (codigo,)):
            return False
        db.ejecutar("""
            INSERT INTO plantas
              (codigo_planta,nombre,tipo,ubicacion,latitud,longitud,
               extension_hectareas,capacidad_mw,fecha_puesta_marcha,
               inversion_inicial,vida_util_anios,estado_operativo,foto_ruta)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            codigo,
            data.get('nombre_planta', ''),
            data.get('tipo_planta', 'solar_fotovoltaica'),
            data.get('ubicacion_planta', ''),
            _lat(data.get('coordenadas_planta', '')),
            _lon(data.get('coordenadas_planta', '')),
            _float(data.get('extension_planta')),
            _float(data.get('capacidad_planta')),
            data.get('fecha_marcha_planta') or None,
            _float(data.get('inversion_planta')),
            _int(data.get('vida_util_planta')),
            'operativa',
            data.get('foto_ruta', ''),
        ))
        return True
    else:
        if any(p['codigo_planta'] == codigo for p in plantas_mem):
            return False
        plantas_mem.append({
            'codigo_planta': codigo, 'nombre': data.get('nombre_planta', ''),
            'tipo': data.get('tipo_planta', ''), 'ubicacion': data.get('ubicacion_planta', ''),
            'capacidad_mw': _float(data.get('capacidad_planta')),
            'estado_operativo': 'operativa',
            'foto_ruta': data.get('foto_ruta', ''),
        })
        return True


def obtener_plantas() -> List[Dict]:
    if db.conectado:
        return db.consultar(
            "SELECT codigo_planta,nombre,tipo,capacidad_mw,estado_operativo FROM plantas ORDER BY nombre")
    return plantas_mem


def eliminar_planta(codigo: str) -> bool:
    if db.conectado:
        db.ejecutar("DELETE FROM plantas WHERE codigo_planta=%s", (codigo,))
    else:
        plantas_mem[:] = [p for p in plantas_mem if p['codigo_planta'] != codigo]
    return True


# ══════════════════════════════════════════════════════════════════
# CRUD — EQUIPOS
# ══════════════════════════════════════════════════════════════════

def agregar_equipo(data: Dict[str, Any]) -> bool:
    n_serie = data.get('n_serie_equipo', '').strip()
    if not n_serie:
        return False
    if db.conectado:
        if db.consultar_uno("SELECT 1 FROM equipos WHERE numero_serie=%s", (n_serie,)):
            return False
        db.ejecutar("""
            INSERT INTO equipos
              (numero_serie,tipo_especifico,marca,modelo,
               potencia_nominal_mw,eficiencia,estado_actual,codigo_planta,foto_ruta)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            n_serie,
            data.get('tipo_equipo', ''),
            data.get('marca_equipo', ''),
            data.get('modelo_equipo', ''),
            _float(data.get('potencia_equipo')),
            _float(data.get('eficiencia_equipo')),
            'operativo',
            data.get('codigo_planta') or None,
            data.get('foto_ruta', ''),
        ))
        return True
    else:
        if any(e['numero_serie'] == n_serie for e in equipos_mem):
            return False
        equipos_mem.append({
            'numero_serie': n_serie, 'tipo_especifico': data.get('tipo_equipo', ''),
            'marca': data.get('marca_equipo', ''),
            'potencia_nominal_mw': _float(data.get('potencia_equipo')),
            'estado_actual': 'operativo',
            'foto_ruta': data.get('foto_ruta', ''),
        })
        return True


def obtener_equipos() -> List[Dict]:
    if db.conectado:
        return db.consultar(
            "SELECT numero_serie,tipo_especifico,marca,potencia_nominal_mw,estado_actual FROM equipos ORDER BY numero_serie")
    return equipos_mem


def eliminar_equipo(n_serie: str) -> bool:
    if db.conectado:
        db.ejecutar("DELETE FROM equipos WHERE numero_serie=%s", (n_serie,))
    else:
        equipos_mem[:] = [e for e in equipos_mem if e['numero_serie'] != n_serie]
    return True


# ══════════════════════════════════════════════════════════════════
# CRUD — ESTACIONES
# ══════════════════════════════════════════════════════════════════

def agregar_estacion(data: Dict[str, Any]) -> bool:
    codigo = data.get('codigo_estacion', '').strip()
    if not codigo:
        return False
    if db.conectado:
        if db.consultar_uno(
                "SELECT 1 FROM estaciones_meteorologicas WHERE codigo_estacion=%s", (codigo,)):
            return False
        eq_raw = data.get('equipos_estacion', '')
        eq_json = json.dumps([x.strip() for x in eq_raw.split(',') if x.strip()])
        db.ejecutar("""
            INSERT INTO estaciones_meteorologicas
              (codigo_estacion,ubicacion,latitud,longitud,
               equipos_instalados,frecuencia_lectura,estado_funcionamiento)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            codigo,
            data.get('ubicacion_estacion', ''),
            _lat(data.get('coordenadas_estacion', '')),
            _lon(data.get('coordenadas_estacion', '')),
            eq_json,
            data.get('frecuencia_estacion', ''),
            'operativa',
        ))
        return True
    else:
        if any(e['codigo_estacion'] == codigo for e in estaciones_mem):
            return False
        estaciones_mem.append({
            'codigo_estacion': codigo, 'ubicacion': data.get('ubicacion_estacion', ''),
            'equipos_instalados': data.get('equipos_estacion', ''),
            'estado_funcionamiento': 'operativa',
        })
        return True


def obtener_estaciones() -> List[Dict]:
    if db.conectado:
        return db.consultar(
            "SELECT codigo_estacion,ubicacion,equipos_instalados,estado_funcionamiento "
            "FROM estaciones_meteorologicas ORDER BY codigo_estacion")
    return estaciones_mem


# ══════════════════════════════════════════════════════════════════
# LECTURAS AVANZADAS
# ══════════════════════════════════════════════════════════════════

def obtener_metricas() -> Dict:
    if db.conectado:
        try:
            c = db.conn.cursor()
            c.execute("CALL sp_dashboard_general(@p,@h,@i,@m)")
            c.execute("SELECT @p,@h,@i,@m")
            row = c.fetchone();
            c.close()
            if row:
                return {'plantas_operativas': row[0] or 0,
                        'produccion_hoy': float(row[1] or 0),
                        'incidencias_abiertas': row[2] or 0,
                        'energia_mes': float(row[3] or 0)}
        except Exception:
            pass
        try:
            def n(q):
                return (db.consultar_uno(q) or {}).get('n', 0)

            return {
                'plantas_operativas': n("SELECT COUNT(*) AS n FROM plantas WHERE estado_operativo='operativa'"),
                'produccion_hoy': float(
                    n("SELECT COALESCE(SUM(energia_generada_mwh),0) AS n FROM produccion_energetica WHERE DATE(fecha_hora)=CURDATE()")),
                'incidencias_abiertas': n("SELECT COUNT(*) AS n FROM incidencias WHERE estado='abierta'"),
                'energia_mes': float(
                    n("SELECT COALESCE(SUM(energia_generada_mwh),0) AS n FROM produccion_energetica WHERE MONTH(fecha_hora)=MONTH(CURDATE())")),
            }
        except Exception:
            pass
    return {'plantas_operativas': len(plantas_mem), 'produccion_hoy': 0,
            'incidencias_abiertas': 0, 'energia_mes': 0}


def obtener_produccion_reciente() -> List[Dict]:
    if db.conectado:
        try:
            return db.consultar("""
                SELECT p.nombre AS planta, pe.potencia_instantanea_mw,
                       pe.energia_generada_mwh, pe.factor_capacidad, pe.fecha_hora
                FROM produccion_energetica pe
                JOIN plantas p ON pe.codigo_planta=p.codigo_planta
                ORDER BY pe.fecha_hora DESC LIMIT 20
            """)
        except Exception:
            pass
    return []


def obtener_incidencias() -> List[Dict]:
    if db.conectado:
        try:
            return db.consultar("""
                SELECT i.codigo_incidencia, p.nombre AS planta, i.tipo_incidencia,
                       i.descripcion, i.impacto_produccion, i.estado, i.fecha_hora
                FROM incidencias i
                JOIN plantas p ON i.codigo_planta=p.codigo_planta
                ORDER BY i.fecha_hora DESC LIMIT 30
            """)
        except Exception:
            pass
    return []


def obtener_mantenimientos() -> List[Dict]:
    if db.conectado:
        try:
            return db.consultar("""
                SELECT m.orden_trabajo, p.nombre AS planta, m.tipo_mantenimiento,
                       m.descripcion_actividades, m.fecha_programada, m.estado
                FROM mantenimientos m
                JOIN plantas p ON m.codigo_planta=p.codigo_planta
                ORDER BY m.fecha_programada DESC LIMIT 30
            """)
        except Exception:
            pass
    return []


# ══════════════════════════════════════════════════════════════════
# GUI — APLICACIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════

class GreenPowerApp:
    def __init__(self, root):
        self.root = root
        self.root.iconbitmap("favicon.ico")
        self.root.title("🌿 GreenPower - Gestión Energía Renovable")
        self.root.geometry("980x740")
        self.root.minsize(820, 620)

        # Estado inicial de la barra
        self._msg_estado = "⚠️ Sin conexión a MySQL — modo offline (datos en memoria)"

        # ✔️ Atributo de plantas en memoria
        self.lista_plantas = []  # ✅ ya está aquí

        self._conectar_db()
        self.crear_menu()
        self.crear_barra_estado()
        self.crear_pestanas()

    def _exportar(self, tipo):
        """Exportar datos a Excel (ejemplo mínimo)"""
        if tipo == "plantas":
            import os
            try:
                import pandas as pd
                # Aquí vendría la lógica real
                print("📊 Exportando plantas a Excel...")
                # Ejemplo basado en self.lista_plantas
                df = pd.DataFrame(
                    self.lista_plantas,
                    columns=["Código", "Nombre", "Tipo", "Capacidad MW", "Estado", "Foto"]
                )
                df.to_excel("Plantas_exportadas.xlsx", index=False)
                messagebox.showinfo("Éxito", "Plantas exportadas a Excel")
            except ImportError:
                messagebox.showerror("Error", "No se encontró pandas. Instala con:\npip install pandas openpyxl")
            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar a Excel:\n{str(e)}")
        else:
            messagebox.showwarning("Atención", f"Exportación no implementada para '{tipo}'")


    def _exportar_pdf(self, tipo):
        """Exportar datos a PDF (ejemplo mínimo)"""
        if tipo == "plantas":
            import os
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
                from reportlab.lib import colors

                doc = SimpleDocTemplate("Plantas_exportadas.pdf", pagesize=letter)
                datos = [["Código", "Nombre", "Tipo", "Capacidad MW", "Estado", "Foto"]] + [
                    list(p) for p in self.lista_plantas
                ]
                tabla = Table(datos)
                tabla.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black)
                ]))

                doc.build([tabla])
                messagebox.showinfo("Éxito", "Plantas exportadas a PDF")
            except ImportError:
                messagebox.showerror("Error", "Instala reportlab:\npip install reportlab")
            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar a PDF:\n{str(e)}")
        else:
            messagebox.showwarning("Atención", f"Exportación PDF no implementada para '{tipo}'")

    def _conectar_db(self):
        global DB_CONFIG, MYSQL_DISPONIBLE

        if not MYSQL_DISPONIBLE:
            self._msg_estado = "⚠️ Instala mysql-connector-python — modo offline"
        elif db.conectar():
            self._msg_estado = (f"✅ Conectado a MySQL  ·  "
                                f"{DB_CONFIG['host']}:{DB_CONFIG['port']}  /  {DB_CONFIG['database']}")
        else:
            self._msg_estado = "⚠️ Sin conexión a MySQL — modo offline (datos en memoria)"

    def reconectar(self):
        try:
            print("🔌 Intentando reconectar a la BD...")
            # Aquí va tu lógica de reconexión real
            print("✅ Base de datos reconectada exitosamente")
            messagebox.showinfo("Éxito", "Base de datos reconectada correctamente")
        except Exception as e:
            print(f"❌ Error al reconectar: {e}")
            messagebox.showerror("Error", f"No se pudo reconectar a la BD:\n{str(e)}")

    def crear_barra_estado(self):
        color = "#27ae60" if db.conectado else "#e67e22"
        pillow_txt = "  |  🖼️ Pillow OK" if PILLOW_DISPONIBLE else "  |  ⚠️ Pillow no instalado"
        self.barra = tk.Label(
            self.root,
            text=self._msg_estado + pillow_txt,
            bg=color, fg="white", anchor='w',
            padx=12, pady=4, font=('Arial', 9))
        self.barra.pack(fill='x', side='bottom')

    def _config_db(self):
        try:
            config_window = tk.Toplevel(self.root)
            config_window.title("⚙️ Configurar Conexión BD")
            config_window.geometry("400x300")
            config_window.resizable(False, False)

            host_var = tk.StringVar(value="localhost")
            port_var = tk.StringVar(value="3306")
            user_var = tk.StringVar(value="root")
            pass_var = tk.StringVar()
            db_var = tk.StringVar(value="greenpower")

            tk.Label(config_window, text="Host:").pack(pady=5)
            tk.Entry(config_window, textvariable=host_var).pack(pady=5)

            tk.Label(config_window, text="Puerto:").pack(pady=5)
            tk.Entry(config_window, textvariable=port_var).pack(pady=5)

            tk.Label(config_window, text="Usuario:").pack(pady=5)
            tk.Entry(config_window, textvariable=user_var).pack(pady=5)

            tk.Label(config_window, text="Contraseña:").pack(pady=5)
            tk.Entry(config_window, textvariable=pass_var, show="*").pack(pady=5)

            tk.Label(config_window, text="Base de datos:").pack(pady=5)
            tk.Entry(config_window, textvariable=db_var).pack(pady=5)

            def guardar_config():
                config = {
                    'host': host_var.get(),
                    'port': int(port_var.get()),
                    'user': user_var.get(),
                    'password': pass_var.get(),
                    'database': db_var.get()
                }
                import json
                with open('config_db.json', 'w') as f:
                    json.dump(config, f)
                config_window.destroy()
                messagebox.showinfo("Éxito", "Configuración guardada")

            tk.Button(config_window, text="Guardar", command=guardar_config,
                      bg="#4CAF50", fg="white").pack(pady=20)

        except Exception as e:
            messagebox.showerror("Error", f"Error en configuración:\n{str(e)}")

    def crear_menu(self):
        mb = tk.Menu(self.root)
        self.root.config(menu=mb)

        arch = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="📁 Archivo", menu=arch)
        arch.add_command(label="🔌 Reconectar BD", command=self.reconectar)
        arch.add_command(label="⚙️ Configurar conexión", command=self._config_db)
        arch.add_separator()
        arch.add_command(label="❌ Salir", command=self.root.quit)

        dat = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="🗄️ Datos", menu=dat)
        dat.add_command(label="🔄 Recargar todo", command=self._recargar_todo)



    def _recargar_todo(self):
        try:
            print("🔄 Recargando todos los datos...")
            self._cargar_plantas()
            self._cargar_equipos()

            if hasattr(self, '_cargar_estaciones'):
                self._cargar_estaciones()
            if hasattr(self, '_cargar_produccion'):
                self._cargar_produccion()
            if hasattr(self, '_cargar_incidencias'):
                self._cargar_incidencias()
            if hasattr(self, '_cargar_mantenimientos'):
                self._cargar_mantenimientos()

            print("✅ Datos recargados correctamente")
            messagebox.showinfo("Éxito", "Todos los datos fueron recargados")

        except Exception as e:
            print(f"❌ Error al recargar: {e}")
            messagebox.showerror("Error", f"No se pudo recargar:\n{str(e)}")

    def crear_pestanas(self):
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill='both', expand=True, padx=10, pady=(10, 0))
        self._tab_plantas()

    def _tabla(self, parent, cols, widths, height=10):
        t = ttk.Treeview(parent, columns=cols, show='headings', height=height)
        for col, w in zip(cols, widths):
            t.heading(col, text=col)
            t.column(col, width=w)
        sb = ttk.Scrollbar(parent, orient='vertical', command=t.yview)
        t.configure(yscroll=sb.set)
        t.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')
        return t

    def _limpiar(self, store: dict):
        for w in store.values():
            if isinstance(w, ttk.Entry):
                w.delete(0, tk.END)
            elif isinstance(w, ttk.Combobox):
                w.current(0)

    def _agregar_planta(self):
        # 1. Tomamos los datos de los campos
        codigo = self.planta_codigo.get().strip()
        nombre = self.planta_nombre.get().strip()
        tipo = self.ep["tipo_planta"].get() if "tipo_planta" in self.ep else ""
        cap = self.ep["capacidad_planta"].get().strip()
        ubic = self.ep["ubicacion_planta"].get().strip()

        # 2. Tomamos la ruta de la foto
        ruta_original = self.img_planta.get_fname()
        print("Ruta original de la foto:", ruta_original)

        # 3. Carpeta de destino
        carpeta_destino = "imagenes_greenpower/plantas"
        os.makedirs(carpeta_destino, exist_ok=True)

        # 4. Nombre de archivo
        nombre_foto = "Sin foto"

        if ruta_original:
            nombre_original = os.path.basename(ruta_original)
            ruta_destino = os.path.join(carpeta_destino, nombre_original)

            print("Ruta origen:", ruta_original)
            print("Ruta destino:", ruta_destino)

            import shutil
            try:
                shutil.copy2(ruta_original, ruta_destino)
                print("✅ Foto copiada correctamente")
                nombre_foto = nombre_original
            except Exception as e:
                print("❌ Error al copiar la foto:", e)
                messagebox.showerror("Error", f"No se pudo guardar la foto:\n{e}")
                nombre_foto = "Sin foto"

        # 5. Validación
        if not codigo or not nombre:
            messagebox.showwarning("Atención", "Código y Nombre son obligatorios")
            return

        # 6. Estado
        estado = "Activa"

        # 7. Guardamos la planta
        self.lista_plantas.append((
            codigo,
            nombre,
            tipo,
            cap or "-",
            estado,
            nombre_foto
        ))

        # 8. Limpiar
        self._limpiar(self.ep)
        self.planta_codigo.set("")
        self.planta_nombre.set("")
        self.img_planta.limpiar()

        # 9. Actualizar tabla
        self._cargar_plantas()

        messagebox.showinfo("Éxito", f"Planta '{nombre}' agregada correctamente")

    def _editar_planta(self):
        messagebox.showinfo("Editar", "Funcionalidad de editar planta")

    def _eliminar_planta(self):
        item = self.tbl_plantas.focus()
        if not item:
            messagebox.showwarning("Atención", "No hay ninguna planta seleccionada")
            return
        valores = self.tbl_plantas.item(item, "values")
        if messagebox.askyesno("Eliminar", f"¿Eliminar planta: {valores[1]}?"):
            self.lista_plantas = [
                p for p in self.lista_plantas
                if p[0] != valores[0]
            ]
            self._cargar_plantas()
            messagebox.showinfo("Éxito", "Planta eliminada")

    def _cargar_plantas(self):
        # Limpiar la tabla antes de volver a llenar
        for i in self.tbl_plantas.get_children():
            self.tbl_plantas.delete(i)
        # Leer lista en memoria
        for p in self.lista_plantas:
            self.tbl_plantas.insert("", "end", values=p)

    def _seleccionar_foto(self, event):
        item = self.tbl_plantas.focus()
        if not item:
            return
        valores = self.tbl_plantas.item(item, "values")

        nombre_foto = valores[5]  # columna "Foto"

        if nombre_foto == "Sin foto":
            self.lbl_miniatura.config(text="Sin foto", image="")
            return

        ruta_foto = f"imagenes_greenpower/plantas/{nombre_foto}"
        import os
        if os.path.exists(ruta_foto):
            from PIL import Image, ImageTk
            img = Image.open(ruta_foto)
            img = img.resize((120, 80), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self.lbl_miniatura.config(image=photo, text="")
            self.lbl_miniatura.image = photo
        else:
            self.lbl_miniatura.config(text="Foto no encontrada", image="")

    def _tab_plantas(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="🏭 Plantas")
        frame.pack(fill='both', expand=True)

        ff = ttk.LabelFrame(frame, text="📝 Nueva / Editar Planta")
        ff.pack(fill='x', padx=10, pady=12)

        # 1. Campos básicos
        self.planta_codigo = tk.StringVar()
        self.planta_nombre = tk.StringVar()

        ttk.Label(ff, text="Código *").grid(row=0, column=0, sticky='w', pady=3)
        ttk.Entry(ff, textvariable=self.planta_codigo, width=22).grid(row=0, column=1, sticky='ew', pady=3)

        ttk.Label(ff, text="Nombre *").grid(row=1, column=0, sticky='w', pady=3)
        ttk.Entry(ff, textvariable=self.planta_nombre, width=22).grid(row=1, column=1, sticky='ew', pady=3)

        # 2. Campos extra
        self.ep = {}

        campos = [
            [('Código *', 'codigo_planta'),
             ('Tipo', 'tipo_planta')],
            [('Nombre *', 'nombre_planta'),
             ('Capacidad (MW)', 'capacidad_planta')],
            [('Ubicación', 'ubicacion_planta'),
             ('Inversión ($)', 'inversion_planta')],
            [('Coordenadas (lat,lon)', 'coordenadas_planta'),
             ('Vida útil (años)', 'vida_util_planta')],
            [('Extensión (ha)', 'extension_planta'),
             ('Fecha marcha (AAAA-MM-DD)', 'fecha_marcha_planta')],
        ]

        combos_p = {
            'tipo_planta': ['solar_fotovoltaica', 'eolica', 'hidroelectrica', 'biomasa']
        }

        for row_idx, par in enumerate(campos, start=2):
            for col_offset, (lbl, key) in enumerate(par):
                c0, c1 = col_offset * 2, col_offset * 2 + 1
                ttk.Label(ff, text=lbl).grid(row=row_idx, column=c0, sticky='w',
                                             pady=3, padx=(0 if c0 == 0 else 14, 6))
                if key in combos_p:
                    w = ttk.Combobox(ff, width=22, state='readonly', values=combos_p[key])
                    w.set(combos_p[key][0])
                else:
                    w = ttk.Entry(ff, width=22)
                w.grid(row=row_idx, column=c1, sticky='ew', pady=3)
                self.ep[key] = w

        ff.columnconfigure(1, weight=1)
        ff.columnconfigure(3, weight=1)

        # 3. Widget imagen
        self.img_planta = SelectorImagen(ff, titulo="🖼️ Foto de la Planta", categoria="planta")
        self.img_planta.grid(row=len(campos) + 2, column=0, columnspan=4, sticky='ew', pady=(10, 4))

        # 4. Botones de entrada
        bf = ttk.Frame(ff)
        bf.grid(row=len(campos) + 3, column=0, columnspan=4, pady=8, sticky='ew')
        ttk.Button(bf, text="➕ Agregar", command=self._agregar_planta).pack(side='left', padx=4)
        ttk.Button(bf, text="✏️ Editar", command=self._editar_planta).pack(side='left', padx=4)
        ttk.Button(bf, text="🗑️ Eliminar selec.", command=self._eliminar_planta).pack(side='left', padx=4)
        ttk.Button(bf, text="🔄 Actualizar tabla", command=self._cargar_plantas).pack(side='left', padx=4)
        ttk.Button(bf, text="📊 Excel", command=lambda: self._exportar('plantas')).pack(side='left', padx=4)
        ttk.Button(bf, text="🖨️ PDF", command=lambda: self._exportar_pdf('plantas')).pack(side='left', padx=4)

        bf.columnconfigure(0, weight=1)

        # 5. Tabla de plantas registradas
        tf = ttk.LabelFrame(frame, text="📋 Plantas Registradas")
        tf.pack(fill='both', expand=True, padx=12, pady=8)

        self.tbl_plantas = self._tabla(
            tf,
            ('Código', 'Nombre', 'Tipo', 'Capacidad MW', 'Estado', 'Foto'),
            (100, 170, 140, 100, 100, 180)
        )

        # 6. Miniatura de la foto (usa tf como padre, en el mismo bloque)
        self.lbl_miniatura = ttk.Label(tf, text="Sin foto")
        self.lbl_miniatura.pack(side='bottom', pady=8)

        # 7. Vincular selección
        self.tbl_plantas.bind("<<TreeviewSelect>>", self._seleccionar_foto)

        # 8. Carga inicial de datos
        self._cargar_plantas()


# ══════════════════════════════════════════════════════════════
# FORMULARIO 2 — EQUIPOS  (con SelectorImagen)
# ══════════════════════════════════════════════════════════════
def _tab_equipos(self):
    frame = ttk.Frame(self.nb)
    self.nb.add(frame, text="⚙️ Equipos")

    ff = ttk.LabelFrame(frame, text="📝 Nuevo Equipo", padding=12)
    ff.pack(fill='x', padx=12, pady=8)
    ff.columnconfigure(1, weight=1)
    ff.columnconfigure(3, weight=1)

    self.ee = {}
    campos_e = [
        [('N° Serie *', 'n_serie_equipo'), ('Tipo específico', 'tipo_equipo')],
        [('Marca', 'marca_equipo'), ('Modelo', 'modelo_equipo')],
        [('Potencia (kW)', 'potencia_equipo'), ('Eficiencia (%)', 'eficiencia_equipo')],
    ]
    for row, par in enumerate(campos_e):
        for col_offset, (lbl, key) in enumerate(par):
            c0, c1 = col_offset * 2, col_offset * 2 + 1
            ttk.Label(ff, text=lbl).grid(row=row, column=c0, sticky='w',
                                         pady=3, padx=(0 if c0 == 0 else 14, 6))
            w = ttk.Entry(ff, width=22)
            w.grid(row=row, column=c1, sticky='ew', pady=3)
            self.ee[key] = w

    # ── Widget imagen Equipo ──────────────────────────────
    self.img_equipo = SelectorImagen(ff, titulo="🖼️ Foto del Equipo", categoria="equipo")
    self.img_equipo.grid(row=3, column=0, columnspan=4, sticky='ew', pady=(10, 4))

    # ── Botones ───────────────────────────────────────────
    bf = ttk.Frame(ff)
    bf.grid(row=4, column=0, columnspan=4, pady=8)
    ttk.Button(bf, text="➕ Agregar", command=self._agregar_equipo).pack(side='left', padx=4)
    ttk.Button(bf, text="🗑️ Eliminar selec.", command=self._eliminar_equipo).pack(side='left', padx=4)
    ttk.Button(bf, text="🔄 Actualizar tabla", command=self._cargar_equipos).pack(side='left', padx=4)
    ttk.Button(bf, text="📊 Excel", command=lambda: self._exportar('equipos')).pack(side='left', padx=4)
    ttk.Button(bf, text="🖨️ PDF", command=lambda: self._exportar_pdf('equipos')).pack(side='left', padx=4)

    # ── Tabla ─────────────────────────────────────────────
    tf = ttk.LabelFrame(frame, text="📋 Equipos Registrados", padding=8)
    tf.pack(fill='both', expand=True, padx=12, pady=8)
    self.tbl_equipos = self._tabla(tf,
                                   ('N° Serie', 'Tipo', 'Marca', 'Potencia MW', 'Estado', 'Foto'),
                                   (130, 130, 120, 100, 100, 180))
    self._cargar_equipos()


# ══════════════════════════════════════════════════════════════
# PESTAÑA ESTACIONES
# ══════════════════════════════════════════════════════════════
def _tab_estaciones(self):
    frame = ttk.Frame(self.nb)
    self.nb.add(frame, text="🌤️ Estaciones")

    ff = ttk.LabelFrame(frame, text="📝 Nueva Estación Meteorológica", padding=12)
    ff.pack(fill='x', padx=12, pady=8)
    ff.columnconfigure(1, weight=1)
    self.est = {}
    for row, (lbl, key) in enumerate([
        ('Código *', 'codigo_estacion'),
        ('Ubicación', 'ubicacion_estacion'),
        ('Coordenadas (lat,lon)', 'coordenadas_estacion'),
        ('Equipos (separados por coma)', 'equipos_estacion'),
        ('Frecuencia', 'frecuencia_estacion'),
    ]):
        ttk.Label(ff, text=lbl).grid(row=row, column=0, sticky='w', pady=3, padx=(0, 10))
        e = ttk.Entry(ff, width=40)
        e.grid(row=row, column=1, sticky='ew', pady=3)
        self.est[key] = e

    bf = ttk.Frame(ff)
    bf.grid(row=6, column=0, columnspan=2, pady=10)
    ttk.Button(bf, text="➕ Agregar", command=self._agregar_estacion).pack(side='left', padx=4)
    ttk.Button(bf, text="🔄 Actualizar tabla", command=self._cargar_estaciones).pack(side='left', padx=4)
    ttk.Button(bf, text="📊 Excel", command=lambda: self._exportar('estaciones')).pack(side='left', padx=4)
    ttk.Button(bf, text="🖨️ PDF", command=lambda: self._exportar_pdf('estaciones')).pack(side='left', padx=4)

    tf = ttk.LabelFrame(frame, text="📋 Estaciones Registradas", padding=8)
    tf.pack(fill='both', expand=True, padx=12, pady=8)
    self.tbl_estaciones = self._tabla(tf,
                                      ('Código', 'Ubicación', 'Equipos', 'Estado'),
                                      (130, 220, 220, 110))
    self._cargar_estaciones()


# ══════════════════════════════════════════════════════════════
# PESTAÑA DASHBOARD
# ══════════════════════════════════════════════════════════════
def _tab_produccion(self):
    frame = ttk.Frame(self.nb)
    self.nb.add(frame, text="📊 Dashboard")

    mf = ttk.LabelFrame(frame, text="📈 Métricas Generales", padding=15)
    mf.pack(fill='x', padx=12, pady=8)

    self.lbl_m = {}
    items = [
        ('plantas_operativas', '🏭 Plantas Operativas:', '—'),
        ('produccion_hoy', '⚡ Producción Hoy (MWh):', '—'),
        ('incidencias_abiertas', '🚨 Incidencias Abiertas:', '—'),
        ('energia_mes', '📅 Energía Mes (MWh):', '—'),
    ]
    for i, (key, lbl, val) in enumerate(items):
        col = (i % 2) * 2
        row = i // 2
        ttk.Label(mf, text=lbl, font=('Arial', 10, 'bold')).grid(
            row=row, column=col, sticky='w', pady=6, padx=(0, 8))
        lv = ttk.Label(mf, text=val, font=('Arial', 11), foreground='#2980b9')
        lv.grid(row=row, column=col + 1, sticky='w', pady=6, padx=(0, 40))
        self.lbl_m[key] = lv

    ttk.Button(mf, text="🔄 Actualizar métricas",
               command=self._cargar_metricas).grid(row=2, column=0, columnspan=4, pady=6)

    tf = ttk.LabelFrame(frame, text="⚡ Producción Reciente", padding=8)
    tf.pack(fill='both', expand=True, padx=12, pady=8)
    self.tbl_prod = self._tabla(tf,
                                ('Planta', 'Potencia MW', 'Energía MWh', 'Factor Cap.', 'Fecha / Hora'),
                                (170, 110, 110, 100, 160))

    self._cargar_metricas()
    self._cargar_produccion()


# ══════════════════════════════════════════════════════════════
# PESTAÑA INCIDENCIAS
# ══════════════════════════════════════════════════════════════
def _tab_incidencias(self):
    frame = ttk.Frame(self.nb)
    self.nb.add(frame, text="🚨 Incidencias")

    ff = ttk.LabelFrame(frame, text="📝 Registrar Incidencia", padding=12)
    ff.pack(fill='x', padx=12, pady=8)
    ff.columnconfigure(1, weight=1)
    self.inc = {}
    for row, (lbl, key) in enumerate([
        ('Código Planta *', 'codigo_planta'),
        ('Tipo incidencia *', 'tipo_incidencia'),
        ('Descripción *', 'descripcion'),
    ]):
        ttk.Label(ff, text=lbl).grid(row=row, column=0, sticky='w', pady=4, padx=(0, 10))
        e = ttk.Entry(ff, width=45)
        e.grid(row=row, column=1, sticky='ew', pady=4)
        self.inc[key] = e

    bf = ttk.Frame(ff)
    bf.grid(row=4, column=0, columnspan=2, pady=8)
    ttk.Button(bf, text="📋 Registrar incidencia", command=self._registrar_incidencia).pack(side='left', padx=4)
    ttk.Button(bf, text="🔄 Recargar lista", command=self._cargar_incidencias).pack(side='left', padx=4)

    tf = ttk.LabelFrame(frame, text="📋 Incidencias Registradas", padding=8)
    tf.pack(fill='both', expand=True, padx=12, pady=8)
    self.tbl_inc = self._tabla(tf,
                               ('Código', 'Planta', 'Tipo', 'Descripción', 'Impacto %', 'Estado', 'Fecha'),
                               (110, 130, 120, 180, 80, 90, 145))
    self._cargar_incidencias()


# ══════════════════════════════════════════════════════════════
# PESTAÑA MANTENIMIENTOS
# ══════════════════════════════════════════════════════════════
def _tab_mantenimientos(self):
    frame = ttk.Frame(self.nb)
    self.nb.add(frame, text="🔧 Mantenimientos")

    ff = ttk.LabelFrame(frame, text="📝 Programar Mantenimiento", padding=12)
    ff.pack(fill='x', padx=12, pady=8)
    ff.columnconfigure(1, weight=1)
    self.mant = {}

    for row, (lbl, key, extra) in enumerate([
        ('Código Planta *', 'codigo_planta', None),
        ('Tipo *', 'tipo', ['preventivo', 'correctivo', 'predictivo']),
        ('Descripción *', 'descripcion', None),
        ('Técnicos (csv)', 'tecnicos', None),
    ]):
        ttk.Label(ff, text=lbl).grid(row=row, column=0, sticky='w', pady=4, padx=(0, 10))
        if extra:
            w = ttk.Combobox(ff, width=20, state='readonly', values=extra)
            w.set(extra[0])
        else:
            w = ttk.Entry(ff, width=50)
        w.grid(row=row, column=1, sticky='ew', pady=4)
        self.mant[key] = w

    bf = ttk.Frame(ff)
    bf.grid(row=5, column=0, columnspan=2, pady=8)
    ttk.Button(bf, text="📋 Programar mantenimiento", command=self._registrar_mantenimiento).pack(side='left', padx=4)
    ttk.Button(bf, text="🔄 Recargar lista", command=self._cargar_mantenimientos).pack(side='left', padx=4)

    tf = ttk.LabelFrame(frame, text="📋 Mantenimientos Programados", padding=8)
    tf.pack(fill='both', expand=True, padx=12, pady=8)
    self.tbl_mant = self._tabla(tf,
                                ('Orden', 'Planta', 'Tipo', 'Descripción', 'Fecha Prog.', 'Estado'),
                                (120, 140, 100, 210, 115, 100))
    self._cargar_mantenimientos()


# ══════════════════════════════════════════════════════════════
# ACCIONES — Plantas
# ══════════════════════════════════════════════════════════════
def _agregar_planta(self):
    try:
        data = {k: (v.get() if hasattr(v, 'get') else '') for k, v in self.ep.items()}
        if not data.get('codigo_planta', '').strip():
            messagebox.showwarning("⚠️", "El Código es obligatorio.")
            return

        # Procesar imagen con Pillow
        codigo = data['codigo_planta'].strip()
        ok_img, resultado_img = self.img_planta.procesar(codigo)
        if not ok_img:
            messagebox.showerror("❌ Imagen", resultado_img)
            return
        data['foto_ruta'] = resultado_img

        if agregar_planta(data):
            msg = f"Planta «{data['nombre_planta']}» guardada."
            if resultado_img:
                msg += f"\n🖼️ Foto guardada en:\n{resultado_img}"
            messagebox.showinfo("✅", msg)
            self.img_planta.limpiar()
            self._limpiar(self.ep)
            self._cargar_plantas()
            self._cargar_metricas()
        else:
            messagebox.showerror("❌", "Código ya existe o datos inválidos.")
    except Exception as ex:
        messagebox.showerror("❌ Error", str(ex))


def _eliminar_planta(self):
    sel = self.tbl_plantas.selection()
    if not sel:
        messagebox.showwarning("⚠️", "Selecciona una planta.")
        return
    cod = str(self.tbl_plantas.item(sel[0])['values'][0])
    if messagebox.askyesno("Confirmar", f"¿Eliminar planta {cod}?"):
        try:
            eliminar_planta(cod)
            self._cargar_plantas()
            self._cargar_metricas()
        except Exception as ex:
            messagebox.showerror("❌", str(ex))


def _editar_planta(self):
    messagebox.showinfo("Editar", "Función editar funcionando")


def _cargar_plantas(self):
    self.tbl_plantas.delete(*self.tbl_plantas.get_children())
    for p in obtener_plantas():
        foto = p.get('foto_ruta', '') or ''
        foto_display = os.path.basename(foto) if foto else '—'
        self.tbl_plantas.insert('', 'end', values=(
            p.get('codigo_planta', ''), p.get('nombre', ''),
            p.get('tipo', ''), p.get('capacidad_mw', ''),
            p.get('estado_operativo', ''), foto_display))


# ── Equipos ───────────────────────────────────────────────────
def _agregar_equipo(self):
    try:
        data = {k: v.get() for k, v in self.ee.items()}
        if not data.get('n_serie_equipo', '').strip():
            messagebox.showwarning("⚠️", "El N° Serie es obligatorio.")
            return

        # Procesar imagen con Pillow
        n_serie = data['n_serie_equipo'].strip()
        ok_img, resultado_img = self.img_equipo.procesar(n_serie)
        if not ok_img:
            messagebox.showerror("❌ Imagen", resultado_img)
            return
        data['foto_ruta'] = resultado_img

        if agregar_equipo(data):
            msg = "Equipo guardado."
            if resultado_img:
                msg += f"\n🖼️ Foto guardada en:\n{resultado_img}"
            messagebox.showinfo("✅", msg)
            self.img_equipo.limpiar()
            self._limpiar(self.ee)
            self._cargar_equipos()
        else:
            messagebox.showerror("❌", "N° Serie ya existe o datos inválidos.")
    except Exception as ex:
        messagebox.showerror("❌", str(ex))


def _eliminar_equipo(self):
    sel = self.tbl_equipos.selection()
    if not sel:
        messagebox.showwarning("⚠️", "Selecciona un equipo.")
        return
    serie = str(self.tbl_equipos.item(sel[0])['values'][0])
    if messagebox.askyesno("Confirmar", f"¿Eliminar equipo {serie}?"):
        try:
            eliminar_equipo(serie)
            self._cargar_equipos()
        except Exception as ex:
            messagebox.showerror("❌", str(ex))


def _cargar_equipos(self):
    self.tbl_equipos.delete(*self.tbl_equipos.get_children())
    for e in obtener_equipos():
        foto = e.get('foto_ruta', '') or ''
        foto_display = os.path.basename(foto) if foto else '—'
        self.tbl_equipos.insert('', 'end', values=(
            e.get('numero_serie', ''), e.get('tipo_especifico', ''),
            e.get('marca', ''), e.get('potencia_nominal_mw', ''),
            e.get('estado_actual', ''), foto_display))


# ── Estaciones ────────────────────────────────────────────────
def _agregar_estacion(self):
    try:
        data = {k: v.get() for k, v in self.est.items()}
        if not data.get('codigo_estacion', '').strip():
            messagebox.showwarning("⚠️", "El Código es obligatorio.")
            return
        if agregar_estacion(data):
            messagebox.showinfo("✅", "Estación guardada.")
            self._limpiar(self.est)
            self._cargar_estaciones()
        else:
            messagebox.showerror("❌", "Código ya existe.")
    except Exception as ex:
        messagebox.showerror("❌", str(ex))


def _cargar_estaciones(self):
    self.tbl_estaciones.delete(*self.tbl_estaciones.get_children())
    for e in obtener_estaciones():
        eq = e.get('equipos_instalados', '')
        if isinstance(eq, (list, dict)): eq = json.dumps(eq, ensure_ascii=False)
        self.tbl_estaciones.insert('', 'end', values=(
            e.get('codigo_estacion', ''), e.get('ubicacion', ''),
            eq, e.get('estado_funcionamiento', '')))


# ── Dashboard ─────────────────────────────────────────────────
def _cargar_metricas(self):
    m = obtener_metricas()
    self.lbl_m['plantas_operativas'].config(text=str(m['plantas_operativas']))
    self.lbl_m['produccion_hoy'].config(text=f"{float(m['produccion_hoy']):.2f}")
    self.lbl_m['incidencias_abiertas'].config(text=str(m['incidencias_abiertas']))
    self.lbl_m['energia_mes'].config(text=f"{float(m['energia_mes']):.2f}")
    self._cargar_produccion()


def _cargar_produccion(self):
    self.tbl_prod.delete(*self.tbl_prod.get_children())
    for r in obtener_produccion_reciente():
        self.tbl_prod.insert('', 'end', values=(
            r.get('planta', ''), r.get('potencia_instantanea_mw', ''),
            r.get('energia_generada_mwh', ''), r.get('factor_capacidad', ''),
            str(r.get('fecha_hora', ''))))


# ── Incidencias ───────────────────────────────────────────────
def _registrar_incidencia(self):
    if not db.conectado:
        messagebox.showwarning("⚠️", "Requiere conexión a MySQL.")
        return
    cod_p = self.inc['codigo_planta'].get().strip()
    tipo = self.inc['tipo_incidencia'].get().strip()
    desc = self.inc['descripcion'].get().strip()
    if not all([cod_p, tipo, desc]):
        messagebox.showwarning("⚠️", "Todos los campos son obligatorios.")
        return
    try:
        c = db.conn.cursor()
        c.execute("CALL sp_crear_incidencia(%s,%s,%s,@cod)", (cod_p, tipo, desc))
        c.execute("SELECT @cod")
        cod = c.fetchone()[0]
        c.close()
        db.conn.commit()
        messagebox.showinfo("✅", f"Incidencia registrada: {cod}")
        self._limpiar(self.inc)
        self._cargar_incidencias()
        self._cargar_metricas()
    except Exception as ex:
        messagebox.showerror("❌", str(ex))


def _cargar_incidencias(self):
    self.tbl_inc.delete(*self.tbl_inc.get_children())
    for r in obtener_incidencias():
        self.tbl_inc.insert('', 'end', values=(
            r.get('codigo_incidencia', ''), r.get('planta', ''),
            r.get('tipo_incidencia', ''), str(r.get('descripcion', ''))[:50],
            r.get('impacto_produccion', ''), r.get('estado', ''),
            str(r.get('fecha_hora', ''))))


# ── Mantenimientos ────────────────────────────────────────────
def _registrar_mantenimiento(self):
    if not db.conectado:
        messagebox.showwarning("⚠️", "Requiere conexión a MySQL.")
        return
    cod_p = self.mant['codigo_planta'].get().strip()
    tipo = self.mant['tipo'].get()
    desc = self.mant['descripcion'].get().strip()
    tecs = json.dumps([t.strip() for t in self.mant['tecnicos'].get().split(',') if t.strip()])
    if not all([cod_p, tipo, desc]):
        messagebox.showwarning("⚠️", "Código, tipo y descripción son obligatorios.")
        return
    try:
        c = db.conn.cursor()
        c.execute("CALL sp_programar_mantenimiento(%s,%s,%s,%s,@orden)", (cod_p, tipo, desc, tecs))
        c.execute("SELECT @orden")
        orden = c.fetchone()[0]
        c.close()
        db.conn.commit()
        messagebox.showinfo("✅", f"Mantenimiento programado: {orden}")
        self._limpiar(self.mant)
        self._cargar_mantenimientos()
    except Exception as ex:
        messagebox.showerror("❌", str(ex))


def _cargar_mantenimientos(self):
    self.tbl_mant.delete(*self.tbl_mant.get_children())
    for r in obtener_mantenimientos():
        self.tbl_mant.insert('', 'end', values=(
            r.get('orden_trabajo', ''), r.get('planta', ''),
            r.get('tipo_mantenimiento', ''), str(r.get('descripcion_actividades', ''))[:50],
            str(r.get('fecha_programada', '')), r.get('estado', '')))


# ── Menú utilidades ───────────────────────────────────────────
def _reconectar(self):
    db.desconectar()
    if db.conectar():
        self.barra.config(
            text=f"✅ Reconectado  ·  {DB_CONFIG['database']}", bg="#27ae60")
        self._recargar_todo()
    else:
        self.barra.config(text="⚠️ Sin conexión — modo offline", bg="#e67e22")


def _config_db(self):
    win = tk.Toplevel(self.root)
    win.title("⚙️ Configurar conexión MySQL")
    win.geometry("360x260")
    win.resizable(False, False)
    entries = {}
    for i, (lbl, key) in enumerate([
        ('Host', 'host'), ('Puerto', 'port'), ('Usuario', 'user'),
        ('Contraseña', 'password'), ('Base de datos', 'database')
    ]):
        ttk.Label(win, text=lbl).grid(row=i, column=0, sticky='w', padx=15, pady=5)
        e = ttk.Entry(win, width=25, show='*' if key == 'password' else '')
        e.insert(0, str(DB_CONFIG.get(key, '')))
        e.grid(row=i, column=1, padx=10, pady=5)
        entries[key] = e

    def guardar():
        for key, e in entries.items():
            val = e.get()
            DB_CONFIG[key] = int(val) if key == 'port' else val
        win.destroy()
        self._reconectar()

    ttk.Button(win, text="💾 Guardar y reconectar", command=guardar).grid(
        row=5, column=0, columnspan=2, pady=15)


def _recargar_todo(self):
    self._cargar_plantas()
    self._cargar_equipos()
    self._cargar_estaciones()
    self._cargar_metricas()
    self._cargar_incidencias()
    self._cargar_mantenimientos()


# ── Exportación ───────────────────────────────────────────────
def _exportar(self, tipo: str):
    dialogo = DialogoFiltros(self.root, tipo)
    self.root.wait_window(dialogo.ventana)
    if not dialogo.resultado: return

    datos, titulo, nombre = self._datos_filtrados(tipo, dialogo.resultado, 'xlsx')
    if not datos:
        messagebox.showwarning("⚠️", f"Sin datos con filtro '{dialogo.resultado['tipo_filtro']}'")
        return
    if Exportador.exportar_excel(datos, titulo, nombre):
        messagebox.showinfo("✅", f"📊 {len(datos)} registros → {nombre}")


def _exportar_pdf(self, tipo: str):
    dialogo = DialogoFiltros(self.root, tipo)
    self.root.wait_window(dialogo.ventana)
    if not dialogo.resultado: return

    datos, titulo, nombre = self._datos_filtrados(tipo, dialogo.resultado, 'pdf')
    if not datos:
        messagebox.showwarning("⚠️", f"Sin datos con filtro '{dialogo.resultado['tipo_filtro']}'")
        return
    if Exportador.exportar_pdf(datos, titulo, nombre):
        messagebox.showinfo("✅", f"🖨️ {len(datos)} registros → {nombre}")


def _datos_filtrados(self, tipo, filtros, ext):
    if tipo == 'plantas':
        datos = self._filtrar_plantas(filtros)
        titulo = "Plantas Filtradas"
        nombre = f"Plantas_{filtros['tipo_filtro']}.{ext}"
    elif tipo == 'equipos':
        datos = self._filtrar_equipos(filtros)
        titulo = "Equipos Filtrados"
        nombre = f"Equipos_{filtros['tipo_filtro']}.{ext}"
    else:
        datos = self._filtrar_estaciones(filtros)
        titulo = "Estaciones Filtradas"
        nombre = f"Estaciones_{filtros['tipo_filtro']}.{ext}"
    return datos, titulo, nombre


def _filtrar_plantas(self, filtros):
    todas, filtradas = obtener_plantas(), []
    for p in todas:
        ok = True
        if filtros['tipo_filtro'] != 'TODOS':
            if filtros['tipo_filtro'] == 'Operativos' and p.get('estado_operativo') != 'operativa':
                ok = False
            elif filtros['tipo_filtro'] == 'No operativos' and p.get('estado_operativo') == 'operativa':
                ok = False
            elif filtros['tipo_filtro'] == 'Solar' and p.get('tipo') != 'solar_fotovoltaica':
                ok = False
            elif filtros['tipo_filtro'] == 'Eólica' and p.get('tipo') != 'eolica':
                ok = False
        if ok: filtradas.append(p)
    return filtradas


def _filtrar_equipos(self, filtros):
    todas, filtradas = obtener_equipos(), []
    for e in todas:
        ok = True
        if filtros['tipo_filtro'] != 'TODOS':
            if filtros['tipo_filtro'] == 'Operativos' and e.get('estado_actual') != 'operativo': ok = False
        if ok: filtradas.append(e)
    return filtradas


def _filtrar_estaciones(self, filtros):
    todas, filtradas = obtener_estaciones(), []
    for e in todas:
        ok = True
        if filtros['tipo_filtro'] != 'TODOS':
            if filtros['tipo_filtro'] == 'Operativas' and e.get('estado_funcionamiento') != 'operativa': ok = False
        if ok: filtradas.append(e)
    return filtradas


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    app = GreenPowerApp(root)
    root.mainloop()
